from collections import defaultdict

from openpyxl import workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.styles.builtins import total
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook import Workbook
import re
import webcolors
from wheel.cli import parse_build_tag


class HtmlTableToExcelConverter:
    CSS_COLOR_NAMES = {
        "black": "000000",
        "silver": "C0C0C0",
        "gray": "808080",
        "white": "FFFFFF",
        "maroon": "800000",
        "red": "FF0000",
        "purple": "800080",
        "fuchsia": "FF00FF",
        "green": "008000",
        "lime": "00FF00",
        "olive": "808000",
        "yellow": "FFFF00",
        "navy": "000080",
        "blue": "0000FF",
        "teal": "008080",
        "aqua": "00FFFF"
    }

    APPLICABLE_CSS_PROPERTIES = {
        'font-weight', 'font-style', 'text-decoration', 'color', 'background-color',
        'text-align', 'vertical-align', 'border', 'border-top', 'border-right', 'border-bottom', 'border-left'
    }

    def __init__(self, table_data):
        self.table_data = table_data
        self.wb = Workbook()
        self.sheet = self.wb.active
        self.current_row = 1
        self.row_heights = defaultdict(list)

    def merge_styles(self, *styles):
        merged = {}
        for style_str in styles:
            if style_str is None:
                style_str = ''
            style_dict = self.parse_style(style_str)
            for key, value in style_dict.items():
                if key in self.APPLICABLE_CSS_PROPERTIES:
                    merged[key] = value
        return merged

    def calculate_row_heigth(self, cell, rowspan, colspan):
        if not cell.value or not cell.alignment or not cell.alignment.wrap_text:
            return

        font_size = cell.font.size or 11
        line_count = self.calculate_line_count(cell, colspan)
        required_height = font_size * line_count + font_size * line_count
        if rowspan == 1:
            self.row_heights[cell.row].append(required_height)

        else:
            per_row_height = required_height / rowspan
            for r in range(cell.row, cell.row + rowspan):
                self.row_heights[r].append(per_row_height)

    def apply_row_heights(self):

        for row_idx, heights in self.row_heights.items():
            if heights:
                required_height = min(heights)
                current_height = self.sheet.row_dimensions[row_idx].height
                if not current_height or current_height < required_height:
                    self.sheet.row_dimensions[row_idx].height = required_height

    def calculate_line_count(self, cell, colspan):
        text = str(cell.value)
        if not text:
            return 1

        total_width = 0
        for col in range(cell.column, cell.column + colspan):
            col_letter = get_column_letter(col)
            col_dim = self.sheet.column_dimensions[col_letter]
            total_width += self.unit_to_px(col_dim.width)


        font_size = cell.font.size or 11
        char_width = font_size * 0.6
        max_chars_per_line = max(1, int(total_width / char_width))
        words = text.split()
        line_count = 1
        current_line_length = 0

        for word in words:
            word_length = len(word)
            if current_line_length > 0:
                word_length += 1

            if current_line_length + word_length < max_chars_per_line:
                current_line_length += word_length
            else:
                if word_length > max_chars_per_line:
                    line_count += (word_length // max_chars_per_line)
                    current_line_length = word_length % max_chars_per_line
                else:
                    line_count += 1
                    current_line_length = word_length
        return max(1, line_count)

    def apply_styles(self, cell, style_dict):
        if not style_dict or not isinstance(style_dict, dict):
            return

        self.apply_alignment(cell, style_dict)
        self.apply_font(cell, style_dict)
        self.apply_background(cell, style_dict)
        self.apply_border(cell, style_dict)

    # Выравнивание
    def apply_alignment(self, cell, style_dict):
        horizontal = style_dict.get('text-align')
        vertical = style_dict.get('vertical-align')

        if vertical == 'middle':
            vertical = 'center'

        alignment = Alignment(
            horizontal=horizontal or 'center',
            vertical=vertical or 'center',
            wrap_text=True
        )
        cell.alignment = alignment

    # Стиль текста
    def apply_font(self, cell, style_dict):
        bold = style_dict.get('font-weight', None)
        italic = style_dict.get('font-style', None)
        underline = style_dict.get('text-decoration', None)

        # Переделываем под формат openpyxl
        if underline:
            underline = 'single' if 'underline' in underline else None

        # Цвет текста
        color = style_dict.get('color') or ""
        color_val = color.strip().lower()
        try:
            if color_val.startswith('#'):
                hex_color = color_val.lstrip('#')
            else:
                hex_color = webcolors.name_to_hex(color_val).lstrip('#')
        except ValueError:
            hex_color = None
        font_color = None
        if hex_color and re.fullmatch(r'[0-9a-fA-F]{6}', hex_color):
            font_color = 'FF' + hex_color.upper()

        # Собираем новый Font
        old_font = cell.font or Font()
        cell.font = Font(
            name=old_font.name,
            size=old_font.size,
            bold=bold if bold is not None else old_font.bold,
            italic=italic if italic is not None else old_font.italic,
            underline=underline if underline is not None else old_font.underline,
            color=font_color if font_color else old_font.color
        )

    # Фон
    def apply_background(self, cell, style_dict):
        # Применение background-color из ячейки
        if 'background-color' in style_dict:
            color = style_dict['background-color']
            try:
                if color.startswith('#'):
                    color = color.lstrip('#')
                else:
                    color = webcolors.name_to_hex(color).lstrip('#')
            except ValueError:
                color = 'FFFFFF'
            end_color = self.expand_short_hex(color)

            if re.fullmatch(r'[0-9a-fA-F]{6}', end_color):
                excel_color = 'FF' + end_color.upper()
                cell.fill = PatternFill(start_color=excel_color, patternType="solid")

        # Добавляем background-color из colgroup для столбца, если он есть и не переопределён в ячейке
        elif 'colgroup' in self.table_data and len(self.table_data['colgroup']) >= cell.column:
            col_style_str = self.table_data['colgroup'][cell.column - 1].get('style', '')
            if col_style_str:
                col_style = self.parse_style(col_style_str)
                bg_color = col_style.get('background-color')
                if bg_color:
                    try:
                        if bg_color.startswith('#'):
                            bg_color_hex = bg_color.lstrip('#')
                        else:
                            bg_color_hex = webcolors.name_to_hex(bg_color).lstrip('#')
                    except ValueError:
                        bg_color_hex = None

                    if bg_color_hex:
                        bg_color_hex = self.expand_short_hex(bg_color_hex)  # Обработка 3-символьного HEX
                        if re.fullmatch(r'[0-9a-fA-F]{6}', bg_color_hex):
                            excel_color = 'FF' + bg_color_hex.upper()
                            cell.fill = PatternFill(start_color=excel_color, patternType="solid")

    # Границы
    def apply_border(self, cell, style_dict):
        def parse_border_side(side_style):
            if not side_style or side_style is None:
                return None
            parts = side_style.split()
            width_px = 1
            line_style = 'thin'
            color = '000000'

            for part in parts:
                if part.endswith('px'):
                    try:
                        width_px = int(part.replace('px', ''))
                    except ValueError:
                        width_px = 1
                elif part in ('solid', 'dashed', 'dotted', 'double', 'groove', \
                              'ridge', 'inset', 'outset'):
                    if part == 'solid':
                        line_style = "thin" if width_px <= 1 \
                            else "medium" if width_px <= 2 else "thick"
                    elif part == 'dashed':
                        line_style = "dashed"
                    elif part == 'dotted':
                        line_style = "dotted"
                    elif part == 'double':
                        line_style = "double"
                    else:
                        line_style = "thin"  # fallback для groove, ridge и др.
                elif part.startswith('#'):
                    color = part.lstrip('#')
                else:
                    try:
                        color = webcolors.name_to_hex(part).lstrip('#')
                    except ValueError:
                        color = "000000"
            color = self.expand_short_hex(color)
            if line_style in ("thin", "medium", "thick"):
                if width_px <= 1:
                    line_style = "thin"
                elif width_px <= 3:
                    line_style = "medium"
                else:
                    line_style = "thick"

            return Side(style=line_style, color=color)

        border_top = parse_border_side(style_dict.get('border-top'))
        border_right = parse_border_side(style_dict.get('border-right'))
        border_bottom = parse_border_side(style_dict.get('border-bottom'))
        border_left = parse_border_side(style_dict.get('border-left'))


        if 'border' in style_dict:
            general_side = parse_border_side(style_dict['border'])
            border_top = border_top if border_top is not None else general_side
            border_right = border_right if border_right is not None else general_side
            border_bottom = border_bottom if border_bottom is not None else general_side
            border_left = border_left if border_left is not None else general_side

        sheet = self.sheet
        merged_ranges = sheet.merged_cells.ranges if sheet.merged_cells else []
        found_merge = False

        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            if min_row <= cell.row <= max_row and min_col <= cell.column <= max_col:
                found_merge = True
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        target_cell = sheet.cell(row=row, column=col)
                        left = border_left if col == min_col else None
                        right = border_right if col == max_col else None
                        top = border_top if row == min_row else None
                        bottom = border_bottom if row == max_row else None
                        target_cell.border = Border(left=left, right=right, top=top, bottom=bottom)
                break

        if not found_merge:
            cell.border = Border(left=border_left, right=border_right, top=border_top, bottom=border_bottom)

    def parse_style(self, style_str):
        style_dict = {}
        if not style_str:
            return style_dict
        for part in style_str.split(';'):
            if ':' in part:
                key, value = part.strip().split(':', 1)
                style_dict[key.strip()] = value.strip()
        return style_dict

    def expand_short_hex(self, hex_color):
        if len(hex_color) == 3:
            return ''.join([c * 2 for c in hex_color])
        elif len(hex_color) == 6:
            return hex_color
        return hex_color

    def set_col_widths(self):
        colgroup = self.table_data.get('colgroup', [])
        table_style = self.parse_style(self.table_data.get('table_style', ''))

        # Общая ширина таблицы
        table_width_px = None
        width_str = table_style.get('width', '')
        if width_str.endswith('px'):
            try:
                table_width_px = int(width_str.replace('px', '').strip())
            except ValueError:
                pass

        section_rows = []
        for section in ['thead', 'tbody', 'tfoot']:
            section_data = self.table_data.get(section, {})
            if isinstance(section_data, dict):
                section_rows.extend(section_data.get('rows', []))

        known_widths = {}
        max_col_index = 0


        # Ширина каждой ячейки
        for row in section_rows:
            cells = row.get('cells', [])
            col_cursor = 0
            for cell in cells:
                colspan = int(cell.get('colspan', 1))
                style = self.parse_style(cell.get('style', ''))
                width_str = style.get('width', '')

                if width_str.endswith('px'):
                    try:
                        px = int(width_str.replace('px', '').strip())
                        per_col_px = px / colspan
                        for offset in range(colspan):
                            col_idx = col_cursor + offset
                            if col_idx not in known_widths:
                                known_widths[col_idx] = per_col_px

                    except ValueError:
                        pass
                elif width_str.endswith('%') and table_width_px is not None:
                    try:
                        percent = int(width_str.replace('%', '').strip())
                        percent = percent / 100
                        per_col_px = table_width_px * percent / colspan
                        for offset in range(colspan):
                            col_idx = col_cursor + offset
                            if col_idx not in known_widths:
                                known_widths[col_idx] = per_col_px
                    except ValueError:
                        pass

                col_cursor += colspan
                max_col_index = max(max_col_index, col_cursor)

        max_length_text_cells = [0] * max_col_index

        for row in section_rows:
            cells = row.get('cells', [])
            col_idx = 0
            for cell in cells:
                colspan = int(cell.get('colspan', 1))
                cell_text = str(cell.get('text', ''))
                length = len(cell_text)
                max_length_text_cells[col_idx] = max(length, max_length_text_cells[col_idx])
                col_idx += colspan


        # Ширина через colgroup
        for col_idx in range(max_col_index):
            if col_idx not in known_widths and col_idx < len(colgroup):
                col_style = self.parse_style(colgroup[col_idx].get('style') or '')
                col_width_str = col_style.get('width', '')
                if col_width_str.endswith('px'):
                    try:
                        px = int(col_width_str.replace('px', '').strip())
                        known_widths[col_idx] = px
                    except ValueError:
                        pass
                elif width_str.endswith('%') and table_width_px is not None:

                    try:
                        percent = int(col_width_str.replace('%', '').strip())
                        percent = percent / 100
                        known_widths[col_idx] = table_width_px * percent
                    except ValueError:
                        pass

        # Назначаем ширину клеткам без явного размера
        unknown_indexes = [i for i in range(max_col_index) if i not in known_widths]
        if table_width_px and unknown_indexes:
            total_known = sum(known_widths.values())
            remaining_px = max(table_width_px - total_known, 0)
            per_col_px = remaining_px / len(unknown_indexes)
            for idx in unknown_indexes:
                known_widths[idx] = per_col_px
        elif not table_width_px and unknown_indexes:
            for idx in unknown_indexes:
                max_length = max_length_text_cells[idx]
                known_widths[idx] = self.auto_width(max_length)

        # Применяем сохраненную ширину
        for col_idx in range(max_col_index):
            px = known_widths.get(col_idx)
            if px:
                excel_width = self.px_to_unit(px)
                col_letter = get_column_letter(col_idx + 1)
                self.sheet.column_dimensions[col_letter].width = excel_width

    def auto_width(self, length):
        style_table =  self.table_data.get('table_style', "")
        style_dict = self.parse_style(style_table)
        weight = style_dict.get("font-size", 11)
        char_width = 0.6
        koef = 1.5
        if length == 0:
            length = 2
        return length * char_width * weight * koef

    def px_to_unit(self, px):
        return (px - 5) / 7

    def unit_to_px(self, unit):
        return unit * 7 + 5

    def add_styles_to_section(self, section):
        section_data = self.table_data[section]
        section_style = section_data.get('style', None)
        rowspan_tracker = {}

        for row in section_data.get('rows', []):
            excel_col_idx = 1

            while (self.current_row, excel_col_idx) in rowspan_tracker:
                rowspan_tracker[(self.current_row, excel_col_idx)] -= 1
                if rowspan_tracker[(self.current_row, excel_col_idx)] == 0:
                    del rowspan_tracker[(self.current_row, excel_col_idx)]
                excel_col_idx += 1

            row_style = row.get('style', None)
            for cell_data in row['cells']:
                colspan = int(cell_data.get('colspan', 1))
                rowspan = int(cell_data.get('rowspan', 1))
                cell = self.sheet.cell(row=self.current_row, column=excel_col_idx)
                cell_style = cell_data.get('style', "")
                cell.value = cell_data.get('text', "")


                if colspan > 1:
                    self.sheet.merge_cells(
                        start_row=self.current_row,
                        start_column=excel_col_idx,
                        end_row=self.current_row,
                        end_column=excel_col_idx + colspan - 1
                    )

                if rowspan > 1:
                    self.sheet.merge_cells(
                        start_row=self.current_row,
                        start_column=excel_col_idx,
                        end_row=self.current_row + rowspan - 1,
                        end_column=excel_col_idx + colspan - 1
                    )

                    for r in range(1, rowspan):
                        for c in range(colspan):
                            key = (self.current_row + r, excel_col_idx + c)
                            rowspan_tracker[key] = rowspan_tracker.get(key, 0) + 1



        # Жирный шрифт по умолчанию
                if section == 'thead':
                    if 'font-weight' not in cell_style:
                        if cell_style:
                            cell_style += "; "
                        cell_style += "font-weight: bold"

                merged_style = self.merge_styles(
                    self.table_data.get('table_style', ""),
                    section_style,
                    row_style,
                    cell_style
                )

                cell = self.get_master_cell(cell)
                self.apply_styles(cell, merged_style)
                self.calculate_row_heigth(cell, rowspan, colspan)
                excel_col_idx += colspan


            self.current_row += 1

    def get_master_cell(self, cell):
        for merged_range in self.sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            if min_row <= cell.row <= max_row and min_col <= cell.column <= max_col:
                return self.sheet.cell(row=min_row, column=min_col)
        return cell

    def convert(self, output_file='test.xlsx'):
        self.set_col_widths()
        self.add_styles_to_section('thead')
        self.add_styles_to_section('tbody')
        self.add_styles_to_section('tfoot')
        self.apply_row_heights()
        self.wb.save(output_file)


        # сделать учет автовысоты по контенту, чтобы текст нормально отображался. параметр fid
        # настройка границ ячейки: цвет, толщина обрамления, стороны(все, одна
        # библиотека конвертации цветов: red -> #f44336
        # поддержка форматирования текста внутри ячейки "<b>Этот</b> текст" чистка inline тегов(b,i) richtext
        # добавление скриптов VBA функция allert showMessage